"""
Business logic utilities for accounting operations.
"""

import io
from datetime import date, timedelta

import openpyxl
from defusedxml import ElementTree as ET
from django.db import models, transaction
from django.utils import timezone
from openpyxl.styles import Border, Font, PatternFill, Side

from .models import (
    AccountingInvoice,
    AccountingPayment,
    BankTransaction,
    ChartOfAccounts,
    Currency,
    DepreciationSchedule,
    Expense,
    FixedAsset,
    LedgerEntry,
)


class DoubleEntryBookkeeping:
    """Handle double-entry bookkeeping operations"""

    @staticmethod
    @transaction.atomic
    def create_journal_entry(
        hospital,
        debit_account_code,
        credit_account_code,
        amount_cents,
        description,
        reference_number,
        transaction_date=None,
        created_by=None,
        source_invoice=None,
        source_payment=None,
        source_expense=None,
        source_payroll=None,
    ):
        """
        Create a double-entry journal entry
        """
        if not transaction_date:
            transaction_date = timezone.now().date()

        # Get accounts
        try:
            debit_account = ChartOfAccounts.objects.get(
                hospital=hospital,
                account_code=debit_account_code,
                is_active=True,
            )
            credit_account = ChartOfAccounts.objects.get(
                hospital=hospital,
                account_code=credit_account_code,
                is_active=True,
            )
        except ChartOfAccounts.DoesNotExist as e:
            raise ValueError(f"Account not found: {e}")

        # Get base currency
        currency = Currency.objects.filter(
            hospital=hospital, is_base_currency=True
        ).first()
        if not currency:
            raise ValueError("No base currency configured")

        # Create ledger entry
        ledger_entry = LedgerEntry.objects.create(
            hospital=hospital,
            transaction_date=transaction_date,
            reference_number=reference_number,
            description=description,
            debit_account=debit_account,
            credit_account=credit_account,
            amount_cents=amount_cents,
            currency=currency,
            created_by=created_by,
            invoice=source_invoice,
            payment=source_payment,
            expense=source_expense,
            payroll=source_payroll,
        )

        return ledger_entry

    @staticmethod
    def post_invoice_entries(invoice):
        """Post accounting entries for an invoice"""
        entries = []

        # Accounts Receivable Dr / Sales Cr
        entries.append(
            DoubleEntryBookkeeping.create_journal_entry(
                hospital=invoice.hospital,
                debit_account_code="1200",  # Accounts Receivable
                credit_account_code="4000",  # Sales Revenue
                amount_cents=invoice.total_cents,
                description=f"Sales invoice {invoice.invoice_number}",
                reference_number=invoice.invoice_number,
                transaction_date=invoice.invoice_date,
                created_by=invoice.created_by,
                source_invoice=invoice,
            )
        )

        # Cost of Sales entries if applicable
        for item in invoice.items.filter(cost_price_cents__gt=0):
            # Cost of Sales Dr / Inventory Cr
            entries.append(
                DoubleEntryBookkeeping.create_journal_entry(
                    hospital=invoice.hospital,
                    debit_account_code="5000",  # Cost of Sales
                    credit_account_code="1300",  # Inventory
                    amount_cents=item.cost_price_cents * int(item.quantity),
                    description=f"Cost of sales for {item.description}",
                    reference_number=invoice.invoice_number,
                    transaction_date=invoice.invoice_date,
                    created_by=invoice.created_by,
                    source_invoice=invoice,
                )
            )

        return entries

    @staticmethod
    def post_payment_entries(payment):
        """Post accounting entries for a payment"""
        # Cash/Bank Dr / Accounts Receivable Cr
        bank_account_code = (
            "1100" if payment.payment_method == "CASH" else "1150"
        )  # noqa: E501

        entry = DoubleEntryBookkeeping.create_journal_entry(
            hospital=payment.hospital,
            debit_account_code=bank_account_code,
            credit_account_code="1200",  # Accounts Receivable
            amount_cents=payment.amount_cents,
            description=f"Payment received - {payment.payment_number}",
            reference_number=payment.payment_number,
            transaction_date=payment.payment_date,
            created_by=payment.received_by,
            source_payment=payment,
        )

        # TDS entry if applicable
        if payment.tds_cents > 0:
            DoubleEntryBookkeeping.create_journal_entry(
                hospital=payment.hospital,
                debit_account_code="1400",  # TDS Receivable
                credit_account_code="2300",  # TDS Payable
                amount_cents=payment.tds_cents,
                description=f"TDS on payment {payment.payment_number}",
                reference_number=payment.payment_number,
                transaction_date=payment.payment_date,
                created_by=payment.received_by,
                source_payment=payment,
            )

        return entry

    @staticmethod
    def post_expense_entries(expense):
        """Post accounting entries for expenses"""
        # Expense Dr / Accounts Payable Cr
        expense_account_code = {
            "MEDICAL_SUPPLIES": "5100",
            "PHARMACEUTICALS": "5110",
            "EQUIPMENT": "5120",
            "UTILITIES": "6100",
            "RENT": "6200",
            "SALARIES": "6300",
            "PROFESSIONAL_FEES": "6400",
            "MAINTENANCE": "6500",
            "INSURANCE": "6600",
            "MARKETING": "6700",
            "ADMINISTRATIVE": "6800",
            "OTHER": "6900",
        }.get(expense.category, "6900")

        entry = DoubleEntryBookkeeping.create_journal_entry(
            hospital=expense.hospital,
            debit_account_code=expense_account_code,
            credit_account_code="2100",  # Accounts Payable
            amount_cents=expense.amount_cents,
            description=f"Expense - {expense.description}",
            reference_number=expense.expense_number,
            transaction_date=expense.expense_date,
            created_by=expense.created_by,
            source_expense=expense,
        )

        return entry

    @staticmethod
    def post_payroll_entries(payroll):
        """Post accounting entries for payroll"""
        entries = []

        # Salary Expense Dr / Salary Payable Cr
        entries.append(
            DoubleEntryBookkeeping.create_journal_entry(
                hospital=payroll.hospital,
                debit_account_code="6300",  # Salary Expense
                credit_account_code="2400",  # Salary Payable
                amount_cents=payroll.net_salary_cents,
                description=f"Salary for {payroll.employee.get_full_name()}",
                reference_number=f"PAY-{payroll.id}",
                transaction_date=payroll.pay_date,
                created_by=payroll.created_by,
                source_payroll=payroll,
            )
        )

        # Employer contributions
        if payroll.pf_employer_cents > 0:
            entries.append(
                DoubleEntryBookkeeping.create_journal_entry(
                    hospital=payroll.hospital,
                    debit_account_code="6310",  # PF Expense
                    credit_account_code="2410",  # PF Payable
                    amount_cents=payroll.pf_employer_cents,
                    description=f"Employer PF contribution for {payroll.employee.get_full_name()}",  # noqa: E501
                    reference_number=f"PAY-{payroll.id}",
                    transaction_date=payroll.pay_date,
                    created_by=payroll.created_by,
                    source_payroll=payroll,
                )
            )

        if payroll.esi_employer_cents > 0:
            entries.append(
                DoubleEntryBookkeeping.create_journal_entry(
                    hospital=payroll.hospital,
                    debit_account_code="6320",  # ESI Expense
                    credit_account_code="2420",  # ESI Payable
                    amount_cents=payroll.esi_employer_cents,
                    description=f"Employer ESI contribution for {payroll.employee.get_full_name()}",  # noqa: E501
                    reference_number=f"PAY-{payroll.id}",
                    transaction_date=payroll.pay_date,
                    created_by=payroll.created_by,
                    source_payroll=payroll,
                )
            )

        return entries


class DepreciationCalculator:
    """Handle asset depreciation calculations"""

    @staticmethod
    def calculate_monthly_depreciation(asset):
        """Calculate monthly depreciation for an asset"""
        annual_depreciation = asset.calculate_annual_depreciation()
        return annual_depreciation // 12

    @staticmethod
    def generate_depreciation_schedule(asset):
        """Generate complete depreciation schedule for an asset"""
        schedule = []
        current_book_value = asset.purchase_cost_cents
        accumulated_depreciation = 0

        for year in range(asset.useful_life_years):
            if asset.depreciation_method == "STRAIGHT_LINE":
                annual_depreciation = (
                    asset.purchase_cost_cents - asset.salvage_value_cents
                ) // asset.useful_life_years
            elif asset.depreciation_method == "REDUCING_BALANCE":
                rate = (
                    asset.depreciation_rate / 100
                    if asset.depreciation_rate
                    else (1 / asset.useful_life_years)
                )
                annual_depreciation = int(current_book_value * rate)
            else:  # DOUBLE_DECLINING
                rate = 2 / asset.useful_life_years
                annual_depreciation = int(current_book_value * rate)

            # Ensure we don't depreciate below salvage value
            if accumulated_depreciation + annual_depreciation > (
                asset.purchase_cost_cents - asset.salvage_value_cents
            ):
                annual_depreciation = (
                    asset.purchase_cost_cents - asset.salvage_value_cents
                ) - accumulated_depreciation

            accumulated_depreciation += annual_depreciation
            current_book_value = (
                asset.purchase_cost_cents - accumulated_depreciation
            )  # noqa: E501

            schedule.append(
                {
                    "year": year + 1,
                    "depreciation_amount": annual_depreciation,
                    "accumulated_depreciation": accumulated_depreciation,
                    "book_value": current_book_value,
                    "date": date(
                        asset.purchase_date.year + year + 1,
                        asset.purchase_date.month,
                        asset.purchase_date.day,
                    ),
                }
            )

            if current_book_value <= asset.salvage_value_cents:
                break

        return schedule

    @staticmethod
    @transaction.atomic
    def process_monthly_depreciation(hospital, processing_date=None):
        """Process depreciation for all assets for a given month"""
        if not processing_date:
            processing_date = timezone.now().date()

        active_assets = FixedAsset.objects.filter(
            hospital=hospital,
            is_active=True,
            purchase_date__lt=processing_date,
        )

        processed_count = 0
        for asset in active_assets:
            # Check if depreciation already processed for this month
            existing_entry = DepreciationSchedule.objects.filter(
                asset=asset,
                depreciation_date__year=processing_date.year,
                depreciation_date__month=processing_date.month,
            ).first()

            if existing_entry:
                continue

            monthly_depreciation = (
                DepreciationCalculator.calculate_monthly_depreciation(asset)
            )
            if monthly_depreciation <= 0:
                continue

            # Create depreciation schedule entry
            new_accumulated = (
                asset.accumulated_depreciation_cents + monthly_depreciation
            )
            new_book_value = asset.purchase_cost_cents - new_accumulated

            DepreciationSchedule.objects.create(
                hospital=hospital,
                asset=asset,
                depreciation_date=processing_date,
                depreciation_amount_cents=monthly_depreciation,
                accumulated_depreciation_cents=new_accumulated,
                book_value_cents=new_book_value,
                is_processed=True,
            )

            # Update asset
            asset.accumulated_depreciation_cents = new_accumulated
            asset.current_book_value_cents = new_book_value
            asset.save(
                update_fields=[
                    "accumulated_depreciation_cents",
                    "current_book_value_cents",
                ]
            )

            # Create accounting entry
            DoubleEntryBookkeeping.create_journal_entry(
                hospital=hospital,
                debit_account_code="6900",  # Depreciation Expense
                credit_account_code="1500",  # Accumulated Depreciation
                amount_cents=monthly_depreciation,
                description=f"Monthly depreciation for {asset.name}",
                reference_number=f"DEP-{asset.asset_code}-{processing_date.strftime('%Y%m')}",  # noqa: E501
                transaction_date=processing_date,
            )

            processed_count += 1

        return processed_count


class TaxCalculator:
    """Handle tax calculations"""

    @staticmethod
    def calculate_gst(amount_cents, gst_rate, is_interstate=False):
        """Calculate GST breakdown"""
        gst_amount = int(amount_cents * gst_rate / 100)

        if is_interstate:
            return {
                "igst_cents": gst_amount,
                "cgst_cents": 0,
                "sgst_cents": 0,
                "total_tax_cents": gst_amount,
            }
        else:
            # Split equally between CGST and SGST
            cgst = gst_amount // 2
            sgst = gst_amount - cgst  # Handle odd amounts
            return {
                "igst_cents": 0,
                "cgst_cents": cgst,
                "sgst_cents": sgst,
                "total_tax_cents": cgst + sgst,
            }

    @staticmethod
    def calculate_tds(gross_amount_cents, tds_rate):
        """Calculate TDS amount"""
        return int(gross_amount_cents * tds_rate / 100)

    @staticmethod
    def get_tax_liability_for_period(
        hospital, start_date, end_date, tax_type="GST"
    ):  # noqa: E501
        """Calculate tax liability for a specific period"""
        # Get all invoices in the period
        invoices = AccountingInvoice.objects.filter(
            hospital=hospital,
            invoice_date__gte=start_date,
            invoice_date__lte=end_date,
            status__in=["PAID", "PARTIAL"],
        )

        total_sales = sum(inv.total_cents for inv in invoices)
        total_tax_collected = sum(inv.tax_cents for inv in invoices)

        # Get input tax credit from expenses
        expenses = Expense.objects.filter(
            hospital=hospital,
            expense_date__gte=start_date,
            expense_date__lte=end_date,
        )

        input_tax_credit = sum(exp.tax_cents for exp in expenses)

        net_tax_liability = total_tax_collected - input_tax_credit

        return {
            "total_sales_cents": total_sales,
            "taxable_sales_cents": total_sales,  # Simplified - all sales are taxable      # noqa: E501
            "tax_collected_cents": total_tax_collected,
            "input_tax_credit_cents": input_tax_credit,
            "net_tax_liability_cents": net_tax_liability,
        }


class ReportGenerator:
    """Generate financial reports"""

    @staticmethod
    def generate_trial_balance(hospital, as_of_date):
        """Generate trial balance report"""
        accounts = ChartOfAccounts.objects.filter(
            hospital=hospital, is_active=True
        )  # noqa: E501
        trial_balance = []

        total_debits = 0
        total_credits = 0

        for account in accounts:
            debit_sum = (
                account.debit_entries.filter(
                    transaction_date__lte=as_of_date
                ).aggregate(total=models.Sum("amount_cents"))["total"]
                or 0
            )

            credit_sum = (
                account.credit_entries.filter(
                    transaction_date__lte=as_of_date
                ).aggregate(total=models.Sum("amount_cents"))["total"]
                or 0
            )

            balance = debit_sum - credit_sum

            if balance != 0:  # Only include accounts with balances
                if balance > 0:
                    debit_balance = balance
                    credit_balance = 0
                    total_debits += debit_balance
                else:
                    debit_balance = 0
                    credit_balance = abs(balance)
                    total_credits += credit_balance

                trial_balance.append(
                    {
                        "account_code": account.account_code,
                        "account_name": account.account_name,
                        "account_type": account.account_type,
                        "debit_balance_cents": debit_balance,
                        "credit_balance_cents": credit_balance,
                    }
                )

        return {
            "accounts": trial_balance,
            "total_debits": total_debits,
            "total_credits": total_credits,
            "is_balanced": total_debits == total_credits,
        }

    @staticmethod
    def generate_profit_loss(hospital, start_date, end_date):
        """Generate Profit & Loss statement"""
        # Income accounts
        income_accounts = ChartOfAccounts.objects.filter(
            hospital=hospital, account_type="INCOME", is_active=True
        )

        total_income = 0
        income_details = []

        for account in income_accounts:
            credit_sum = (
                account.credit_entries.filter(
                    transaction_date__gte=start_date,
                    transaction_date__lte=end_date,
                ).aggregate(total=models.Sum("amount_cents"))["total"]
                or 0
            )

            debit_sum = (
                account.debit_entries.filter(
                    transaction_date__gte=start_date,
                    transaction_date__lte=end_date,
                ).aggregate(total=models.Sum("amount_cents"))["total"]
                or 0
            )

            net_income = credit_sum - debit_sum
            total_income += net_income

            if net_income != 0:
                income_details.append(
                    {
                        "account_code": account.account_code,
                        "account_name": account.account_name,
                        "amount_cents": net_income,
                    }
                )

        # Expense accounts
        expense_accounts = ChartOfAccounts.objects.filter(
            hospital=hospital, account_type="EXPENSES", is_active=True
        )

        total_expenses = 0
        expense_details = []

        for account in expense_accounts:
            debit_sum = (
                account.debit_entries.filter(
                    transaction_date__gte=start_date,
                    transaction_date__lte=end_date,
                ).aggregate(total=models.Sum("amount_cents"))["total"]
                or 0
            )

            credit_sum = (
                account.credit_entries.filter(
                    transaction_date__gte=start_date,
                    transaction_date__lte=end_date,
                ).aggregate(total=models.Sum("amount_cents"))["total"]
                or 0
            )

            net_expense = debit_sum - credit_sum
            total_expenses += net_expense

            if net_expense != 0:
                expense_details.append(
                    {
                        "account_code": account.account_code,
                        "account_name": account.account_name,
                        "amount_cents": net_expense,
                    }
                )

        net_profit = total_income - total_expenses

        return {
            "period": f"{start_date} to {end_date}",
            "income": income_details,
            "expenses": expense_details,
            "total_income_cents": total_income,
            "total_expenses_cents": total_expenses,
            "net_profit_cents": net_profit,
        }

    @staticmethod
    def generate_balance_sheet(hospital, as_of_date):
        """Generate Balance Sheet"""
        balance_sheet = {
            "as_of_date": as_of_date,
            "assets": {
                "current_assets": [],
                "fixed_assets": [],
                "total_assets_cents": 0,
            },
            "liabilities": {
                "current_liabilities": [],
                "long_term_liabilities": [],
                "total_liabilities_cents": 0,
            },
            "equity": {"equity_accounts": [], "total_equity_cents": 0},
        }

        # Assets
        asset_accounts = ChartOfAccounts.objects.filter(
            hospital=hospital, account_type="ASSETS", is_active=True
        )

        for account in asset_accounts:
            balance = account.balance
            if balance > 0:
                asset_data = {
                    "account_code": account.account_code,
                    "account_name": account.account_name,
                    "amount_cents": balance,
                }

                if account.account_subtype == "CURRENT_ASSETS":
                    balance_sheet["assets"]["current_assets"].append(
                        asset_data
                    )  # noqa: E501
                else:
                    balance_sheet["assets"]["fixed_assets"].append(asset_data)

                balance_sheet["assets"]["total_assets_cents"] += balance

        # Liabilities
        liability_accounts = ChartOfAccounts.objects.filter(
            hospital=hospital, account_type="LIABILITIES", is_active=True
        )

        for account in liability_accounts:
            balance = account.balance
            if balance > 0:
                liability_data = {
                    "account_code": account.account_code,
                    "account_name": account.account_name,
                    "amount_cents": balance,
                }

                if account.account_subtype == "CURRENT_LIABILITIES":
                    balance_sheet["liabilities"]["current_liabilities"].append(
                        liability_data
                    )
                else:
                    balance_sheet["liabilities"][
                        "long_term_liabilities"
                    ].append(  # noqa: E501
                        liability_data
                    )

                balance_sheet["liabilities"][
                    "total_liabilities_cents"
                ] += balance  # noqa: E501

        # Equity
        equity_accounts = ChartOfAccounts.objects.filter(
            hospital=hospital, account_type="EQUITY", is_active=True
        )

        for account in equity_accounts:
            balance = account.balance
            if balance != 0:
                balance_sheet["equity"]["equity_accounts"].append(
                    {
                        "account_code": account.account_code,
                        "account_name": account.account_name,
                        "amount_cents": balance,
                    }
                )
                balance_sheet["equity"]["total_equity_cents"] += balance

        return balance_sheet


class ExportEngine:
    """Handle various export formats"""

    @staticmethod
    def export_to_excel(data, headers, filename):
        """Export data to Excel format"""
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin"),
        )

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        # Write data
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Save to BytesIO
        excel_buffer = io.BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)

        return excel_buffer

    @staticmethod
    def export_to_tally_xml(transactions, export_type="SALES"):
        """Export to Tally-compatible XML format"""
        root = ET.Element("ENVELOPE")
        header = ET.SubElement(root, "HEADER")
        ET.SubElement(header, "TALLYREQUEST").text = "Import Data"

        body = ET.SubElement(root, "BODY")
        import_data = ET.SubElement(body, "IMPORTDATA")
        request_desc = ET.SubElement(import_data, "REQUESTDESC")
        ET.SubElement(request_desc, "REPORTNAME").text = "Vouchers"

        request_data = ET.SubElement(import_data, "REQUESTDATA")

        for trans in transactions:
            voucher = ET.SubElement(request_data, "TALLYMESSAGE")
            voucher_data = ET.SubElement(voucher, "VOUCHER")

            if export_type == "SALES":
                ET.SubElement(voucher_data, "VOUCHERTYPENAME").text = "Sales"
                ET.SubElement(voucher_data, "PARTYLEDGERNAME").text = transaction.get(
                    "party_name", ""
                )  # noqa: E501  # noqa: E501  # noqa: E501   # noqa: E501
                ET.SubElement(voucher_data, "DATE").text = transaction.get(
                    "date", ""
                ).strftime("%Y%m%d")
                ET.SubElement(voucher_data, "AMOUNT").text = str(
                    transaction.get("amount_cents", 0) / 100
                )
            elif export_type == "PURCHASE":
                ET.SubElement(
                    voucher_data, "VOUCHERTYPENAME"
                ).text = "Purchase"  # noqa: E501
            # Add other voucher types as needed

        return ET.tostring(root, encoding="unicode")

    @staticmethod
    def export_gst_returns(
        hospital, start_date, end_date, return_type="GSTR1"
    ):  # noqa: E501
        """Export GST returns in JSON format"""
        if return_type == "GSTR1":
            # B2B supplies
            b2b_supplies = []
            invoices = AccountingInvoice.objects.filter(
                hospital=hospital,
                invoice_date__gte=start_date,
                invoice_date__lte=end_date,
                invoice_type="CORPORATE",
            )

            for invoice in invoices:
                if invoice.customer and invoice.customer.gstin:
                    b2b_supplies.append(
                        {
                            "ctin": invoice.customer.gstin,
                            "inv": [
                                {
                                    "inum": invoice.invoice_number,
                                    "idt": invoice.invoice_date.strftime(
                                        "%d-%m-%Y"
                                    ),  # noqa: E501
                                    "val": invoice.total_cents / 100,
                                    "pos": "07",  # Place of supply - needs to be configured      # noqa: E501
                                    "rchrg": "N",
                                    "inv_typ": "R",
                                    "itms": [],
                                }
                            ],
                        }
                    )

            return {
                "gstin": "",  # Hospital's GSTIN - needs to be configured
                "ret_period": start_date.strftime("%m%Y"),
                "b2b": b2b_supplies,
            }

        # Add other GST return types as needed
        return {}


class AgeingReportGenerator:
    """Generate ageing reports for receivables and payables"""

    @staticmethod
    def generate_receivables_ageing(hospital, as_of_date):
        """Generate ageing report for accounts receivable"""
        unpaid_invoices = AccountingInvoice.objects.filter(
            hospital=hospital,
            status__in=["SENT", "OVERDUE", "PARTIAL"],
            balance_cents__gt=0,
        )

        ageing_buckets = {"0-30": [], "31-60": [], "61-90": [], "90+": []}

        for invoice in unpaid_invoices:
            days_outstanding = (as_of_date - invoice.due_date).days

            ageing_data = {
                "invoice_number": invoice.invoice_number,
                "customer": str(invoice.customer or invoice.patient),
                "invoice_date": invoice.invoice_date,
                "due_date": invoice.due_date,
                "total_cents": invoice.total_cents,
                "balance_cents": invoice.balance_cents,
                "days_outstanding": days_outstanding,
            }

            if days_outstanding <= 30:
                ageing_buckets["0-30"].append(ageing_data)
            elif days_outstanding <= 60:
                ageing_buckets["31-60"].append(ageing_data)
            elif days_outstanding <= 90:
                ageing_buckets["61-90"].append(ageing_data)
            else:
                ageing_buckets["90+"].append(ageing_data)

        # Calculate totals for each bucket
        bucket_totals = {}
        for bucket, invoices in ageing_buckets.items():
            bucket_totals[bucket] = sum(
                inv["balance_cents"] for inv in invoices
            )  # noqa: E501

        return {
            "as_of_date": as_of_date,
            "ageing_buckets": ageing_buckets,
            "bucket_totals": bucket_totals,
            "grand_total": sum(bucket_totals.values()),
        }


class BankReconciliationHelper:
    """Help with bank reconciliation"""

    @staticmethod
    def auto_match_transactions(bank_account, tolerance_cents=100):
        """Auto-match bank transactions with payments/expenses"""
        unreconciled_bank_txns = BankTransaction.objects.filter(
            bank_account=bank_account, is_reconciled=False
        )

        matched_count = 0

        for bank_txn in unreconciled_bank_txns:
            if bank_txn.transaction_type == "CREDIT":
                # Try to match with payments
                matching_payments = AccountingPayment.objects.filter(
                    bank_account=bank_account,
                    payment_date=bank_txn.transaction_date,
                    amount_cents__gte=bank_txn.amount_cents - tolerance_cents,
                    amount_cents__lte=bank_txn.amount_cents + tolerance_cents,
                    status="CLEARED",
                ).exclude(banktransaction__isnull=False)

                if matching_payments.count() == 1:
                    payment = matching_payments.first()
                    bank_txn.reconciled_payment = payment
                    bank_txn.is_reconciled = True
                    bank_txn.reconciled_at = timezone.now()
                    bank_txn.save()
                    matched_count += 1

            else:  # DEBIT
                # Try to match with expenses
                matching_expenses = Expense.objects.filter(
                    payment_date=bank_txn.transaction_date,
                    net_amount_cents__gte=bank_txn.amount_cents
                    - tolerance_cents,  # noqa: E501
                    net_amount_cents__lte=bank_txn.amount_cents
                    + tolerance_cents,  # noqa: E501
                    is_paid=True,
                ).exclude(banktransaction__isnull=False)

                if matching_expenses.count() == 1:
                    expense = matching_expenses.first()
                    bank_txn.reconciled_expense = expense
                    bank_txn.is_reconciled = True
                    bank_txn.reconciled_at = timezone.now()
                    bank_txn.save()
                    matched_count += 1

        return matched_count


class ComplianceReporter:
    """Generate compliance reports"""

    @staticmethod
    def generate_tds_return(hospital, quarter, financial_year):
        """Generate TDS return data"""
        from .models import TDSEntry

        # Calculate quarter dates
        quarter_months = {
            "Q1": (4, 6),  # Apr-Jun
            "Q2": (7, 9),  # Jul-Sep
            "Q3": (10, 12),  # Oct-Dec
            "Q4": (1, 3),  # Jan-Mar
        }

        start_month, end_month = quarter_months[quarter]
        if quarter == "Q4":
            start_date = date(financial_year + 1, start_month, 1)
            end_date = date(financial_year + 1, end_month + 1, 1) - timedelta(
                days=1
            )  # noqa: E501
        else:
            start_date = date(financial_year, start_month, 1)
            end_date = date(financial_year, end_month + 1, 1) - timedelta(
                days=1
            )  # noqa: E501

        tds_entries = TDSEntry.objects.filter(
            hospital=hospital,
            deduction_date__gte=start_date,
            deduction_date__lte=end_date,
        )

        # Group by section and vendor/employee
        tds_summary = {}
        for entry in tds_entries:
            key = (entry.section, entry.vendor_id or entry.employee_id)
            if key not in tds_summary:
                tds_summary[key] = {
                    "section": entry.section,
                    "deductee": str(entry.vendor or entry.employee),
                    "pan": getattr(entry.vendor, "pan", "")
                    or getattr(entry.employee, "profile", {}).get("pan", ""),
                    "gross_amount_cents": 0,
                    "tds_amount_cents": 0,
                }

            tds_summary[key]["gross_amount_cents"] += entry.gross_amount_cents
            tds_summary[key]["tds_amount_cents"] += entry.tds_amount_cents

        return {
            "quarter": quarter,
            "financial_year": financial_year,
            "start_date": start_date,
            "end_date": end_date,
            "tds_entries": list(tds_summary.values()),
            "total_tds_cents": sum(
                entry["tds_amount_cents"] for entry in tds_summary.values()
            ),
        }
